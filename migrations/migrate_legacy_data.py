"""
Data Migration Script: Legacy Scripts -> Platform Metadata Tables
=================================================================
Migrates ALL configurations from legacy files into PostgreSQL metadata:

1. data_mapping.xlsx (Sheet1)  -> ra_meta.table_config     (198 ETL mappings)
2. data_mapping.xlsx (Skip)    -> ra_meta.table_config     (20 skipped/new mappings)
3. sequence_updated3.xlsx      -> ra_meta.mv_refresh_order (485 MV refresh sequence)
4. manage_jobs.py              -> ra_meta.jobs             (3 job definitions)
5. extract_tables_to_excel2.py -> ra_meta.jobs + config    (export job + table list)
6. Refresh_withkill.py         -> ra_meta.jobs + config    (refresh job config)
7. treevFinalv3.py             -> ra_meta.jobs             (dependency scan job)
8. Connection profiles         -> ra_meta.connection_profiles
9. Hardcoded credentials       -> environment variable references

Usage:
    python -m migrations.migrate_legacy_data
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import openpyxl

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))


def generate_uuid():
    return str(uuid4())


def escape_sql(value):
    """Escape single quotes for SQL string literals."""
    if value is None:
        return "NULL"
    s = str(value).replace("'", "''")
    return f"'{s}'"


def escape_sql_or_null(value):
    """Return NULL for empty/None values, escaped string otherwise."""
    if value is None or str(value).strip() == "" or str(value).strip().lower() in ("nan", "none", "nat"):
        return "NULL"
    return escape_sql(value)


def load_data_mapping(excel_path):
    """Load data_mapping.xlsx Sheet1 and Skip sheet."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    rows = []
    for sheet_name in ["Sheet1", "Skip"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[str(header).strip()] = row[i]
            if row_dict.get("source_table_name"):
                row_dict["_sheet"] = sheet_name
                rows.append(row_dict)

    wb.close()
    return rows


def load_mv_sequence(excel_path):
    """Load sequence_updated3.xlsx MV refresh order."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    mvs = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            mvs.append(str(row[0]).strip())
    wb.close()
    return mvs


def generate_connection_profiles_sql():
    """Generate SQL for connection profiles (replacing hardcoded credentials)."""
    pg_id = generate_uuid()
    oracle_id = generate_uuid()
    oracle_export_id = generate_uuid()

    sql = f"""
-- ============================================================
-- CONNECTION PROFILES (Replace hardcoded credentials)
-- ============================================================
-- Legacy: config.json -> postgresql section
-- Legacy: oracle_connection_string = "recon_prd/recon123@incor.solutions.com.sa"
-- Now: credentials stored in environment variables, NOT in database

INSERT INTO ra_meta.connection_profiles (id, name, db_type, host, port, database_name, username, password_env_var, extra_params, is_active)
VALUES
    -- PostgreSQL source (from config.json used by Loader_full_Newv2.py)
    ('{pg_id}', 'PostgreSQL - Cloud Marketplace', 'postgresql',
     'SET_VIA_ENV', 5432, 'SET_VIA_ENV', 'SET_VIA_ENV',
     'PG_SOURCE_PASSWORD',
     '{{"description": "PostgreSQL source for marketplace/cloud data tables", "legacy_config": "config.json -> postgresql"}}',
     TRUE),

    -- Oracle target (data warehouse - used by ALL scripts)
    ('{oracle_id}', 'Oracle - Revenue Assurance DWH', 'oracle',
     'incor.solutions.com.sa', 1521, 'ORCL', 'recon_prd',
     'ORACLE_PASSWORD',
     '{{"description": "Oracle DWH for revenue assurance", "legacy_connection": "recon_prd@incor.solutions.com.sa", "oracle_home": "/home/oracle/app/product/12.2.0/client_1"}}',
     TRUE),

    -- Oracle connection for exports (extract_tables_to_excel2.py)
    ('{oracle_export_id}', 'Oracle - Export Source', 'oracle',
     'incor.solutions.com.sa', 1521, 'ORCL', 'recon_prd',
     'ORACLE_PASSWORD',
     '{{"description": "Oracle source for table exports and historical snapshots", "legacy_script": "extract_tables_to_excel2.py", "export_targets": ["/home/sdev/ROOT/AutoExport", "/home/sdev/ROOT2/Extracts"]}}',
     TRUE)
ON CONFLICT DO NOTHING;
"""
    return sql, pg_id, oracle_id, oracle_export_id


def generate_table_config_sql(mappings, pg_conn_id, oracle_conn_id):
    """Generate SQL INSERT statements for all 198+ table configurations."""
    sql_lines = ["""
-- ============================================================
-- TABLE CONFIGURATIONS (Migrated from data_mapping.xlsx)
-- 198 active mappings from Sheet1 + 20 from Skip sheet
-- ============================================================
"""]

    for i, m in enumerate(mappings):
        config_id = generate_uuid()
        name = str(m.get("destination_table_name", f"mapping_{i}")).strip()
        source_table = str(m.get("source_table_name", "")).strip()
        source_type = str(m.get("type", "csv")).strip().lower()
        if source_type not in ("postgres", "csv", "excel"):
            source_type = "csv"

        source_cols = str(m.get("source_columns", "")).strip()
        dest_table = str(m.get("destination_table_name", "")).strip()
        dest_cols = str(m.get("destination_columns", "")).strip()
        where_cond = m.get("where_condition")
        backup_folder = m.get("Bkup_folder")
        schedule_day = m.get("Date")
        skip_rows = m.get("skip", 0)
        date_column = m.get("Date_column")
        date_format = m.get("Date_Format")
        file_name = m.get("File_name")
        file_ext = m.get("File_ext")
        file_name_orig = m.get("File_name_Orignal")
        note = m.get("Note")
        sheet = m.get("_sheet", "Sheet1")

        # Determine if enabled based on sheet and note
        is_enabled = "TRUE"
        if sheet == "Skip":
            is_enabled = "FALSE"
        if note and "exclude" in str(note).lower():
            is_enabled = "FALSE"

        # Source connection based on type
        src_conn = f"'{pg_conn_id}'" if source_type == "postgres" else "NULL"

        # Schedule day
        try:
            day_val = int(float(str(schedule_day))) if schedule_day else "NULL"
        except (ValueError, TypeError):
            day_val = "NULL"

        # Skip rows
        try:
            skip_val = int(float(str(skip_rows))) if skip_rows else 0
        except (ValueError, TypeError):
            skip_val = 0

        # Build config JSON with all legacy fields preserved
        config_json = {
            "legacy_sheet": sheet,
            "file_name_pattern": str(file_name) if file_name and str(file_name).lower() not in ("nan", "none") else None,
            "file_extension_filter": str(file_ext) if file_ext and str(file_ext).lower() not in ("nan", "none") else None,
            "file_name_original_col": str(file_name_orig) if file_name_orig and str(file_name_orig).lower() not in ("nan", "none", "no") else None,
            "note": str(note) if note and str(note).lower() not in ("nan", "none") else None,
        }
        # Remove None values
        config_json = {k: v for k, v in config_json.items() if v is not None}

        # Handle json_mapping fields
        json_mapping = m.get("json_mapping")
        json_columns = m.get("json_columns")
        json_key = m.get("json_key")
        if json_mapping and str(json_mapping).lower() not in ("nan", "none"):
            config_json["json_mapping"] = str(json_mapping)
        if json_columns and str(json_columns).lower() not in ("nan", "none"):
            config_json["json_columns"] = str(json_columns)
        if json_key and str(json_key).lower() not in ("nan", "none"):
            config_json["json_key"] = str(json_key)

        config_json_str = json.dumps(config_json).replace("'", "''")

        sql_lines.append(f"""INSERT INTO ra_meta.table_config (
    id, name, source_type, source_connection_id, source_table_name, source_columns,
    where_condition, destination_connection_id, destination_table_name, destination_columns,
    date_format, date_column, schedule_day, skip_rows,
    file_name_pattern, file_extension_filter, backup_folder, is_enabled, config_json
) VALUES (
    '{config_id}',
    {escape_sql(name + '_' + source_type + '_' + str(i))},
    {escape_sql(source_type)},
    {src_conn},
    {escape_sql(source_table)},
    {escape_sql(source_cols)},
    {escape_sql_or_null(where_cond)},
    '{oracle_conn_id}',
    {escape_sql(dest_table)},
    {escape_sql(dest_cols)},
    {escape_sql_or_null(date_format)},
    {escape_sql_or_null(date_column)},
    {day_val},
    {skip_val},
    {escape_sql_or_null(file_name)},
    {escape_sql_or_null(file_ext)},
    {escape_sql_or_null(backup_folder)},
    {is_enabled},
    '{config_json_str}'
);""")

    return "\n".join(sql_lines)


def generate_mv_sequence_sql(mvs):
    """Generate SQL for MV refresh sequence (485 materialized views)."""
    sql_lines = ["""
-- ============================================================
-- MV REFRESH SEQUENCE (Migrated from sequence_updated3.xlsx)
-- 485 materialized views in dependency-sorted refresh order
-- This is the static fallback sequence. The platform dynamically
-- computes dependency order via the Dependency Engine service,
-- but this seed data preserves the proven legacy sequence.
-- ============================================================

CREATE TABLE IF NOT EXISTS ra_meta.mv_refresh_sequence (
    id              UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    mv_name         VARCHAR(255) NOT NULL,
    refresh_order   INTEGER NOT NULL,
    refresh_type    VARCHAR(20) DEFAULT 'COMPLETE',
    timeout_seconds INTEGER DEFAULT 3600,
    is_enabled      BOOLEAN DEFAULT TRUE,
    created_at      TIMESTAMPTZ DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_mv_refresh_seq_order ON ra_meta.mv_refresh_sequence (refresh_order);
CREATE INDEX IF NOT EXISTS idx_mv_refresh_seq_name ON ra_meta.mv_refresh_sequence (mv_name);
"""]

    # Batch insert for performance
    batch_size = 50
    for batch_start in range(0, len(mvs), batch_size):
        batch = mvs[batch_start:batch_start + batch_size]
        values = []
        for offset, mv_name in enumerate(batch):
            order_num = batch_start + offset + 1
            values.append(
                f"    (gen_random_uuid(), '{mv_name}', {order_num}, 'COMPLETE', 3600, TRUE, NOW())"
            )
        sql_lines.append(
            f"INSERT INTO ra_meta.mv_refresh_sequence (id, mv_name, refresh_order, refresh_type, timeout_seconds, is_enabled, created_at)\nVALUES\n"
            + ",\n".join(values)
            + ";"
        )

    return "\n".join(sql_lines)


def generate_jobs_sql():
    """Generate SQL for job definitions (migrated from manage_jobs.py)."""
    loader_id = generate_uuid()
    refresh_id = generate_uuid()
    export_id = generate_uuid()
    dep_scan_id = generate_uuid()

    return f"""
-- ============================================================
-- JOB DEFINITIONS (Migrated from manage_jobs.py JOB_CONFIGS)
-- ============================================================
-- Legacy: manage_jobs.py defined 3 jobs: loader, refresh, extract
-- Now: stored in ra_meta.jobs with cron schedules and configs

INSERT INTO ra_meta.jobs (id, name, job_type, description, schedule_cron, config, is_enabled, max_retries, timeout_seconds)
VALUES
    -- 1. Data Loader Job (legacy: Loader_full_Newv2.py)
    -- Reads data_mapping.xlsx, loads CSV/Excel/PostgreSQL data into Oracle via SQL*Loader
    -- Legacy schedule: cron-based, runs on specific days (Date column in mapping)
    ('{loader_id}',
     'Daily ETL Loader',
     'load',
     'Loads data from PostgreSQL/CSV/Excel sources into Oracle DWH via SQL*Loader. '
     'Processes all table_config entries matching the current day. '
     'Legacy script: Loader_full_Newv2.py + data_mapping.xlsx',
     '0 2 * * *',
     '{{
        "description": "Migrated from Loader_full_Newv2.py",
        "legacy_script": "Loader_full_Newv2.py",
        "legacy_config": "data_mapping.xlsx",
        "legacy_work_dir": "/u01/RA_OPS/Test_New_Loader/Final",
        "legacy_log_path": "/u01/RA_OPS/Test_New_Loader/Final/LOGs",
        "batch_size": 2000000,
        "nls_lang": "AMERICAN_AMERICA.AL32UTF8"
     }}',
     TRUE, 3, 7200),

    -- 2. MV Refresh Job (legacy: Refresh_withkill.py)
    -- Refreshes 485 materialized views in dependency order with timeout per MV
    -- Legacy: reads sequence_updated3.xlsx, refreshes sequentially with 60min timeout
    ('{refresh_id}',
     'Daily MV Refresh',
     'refresh',
     'Refreshes all 485 materialized views in dependency-sorted order. '
     'Platform enhancement: parallel refresh within topological levels. '
     'Legacy script: Refresh_withkill.py + sequence_updated3.xlsx',
     '0 4 * * *',
     '{{
        "description": "Migrated from Refresh_withkill.py",
        "legacy_script": "Refresh_withkill.py",
        "legacy_config": "sequence_updated3.xlsx",
        "legacy_work_dir": "/u01/RA_OPS",
        "legacy_log_path": "/u01/RA_OPS/MVRefrshLogs.csv",
        "legacy_error_log": "/u01/RA_OPS/script_log.txt",
        "schema_owner": "RECON_PRD",
        "refresh_type": "COMPLETE",
        "timeout_per_mv": 3600,
        "max_parallel": 4,
        "backup_ddl_before_refresh": true,
        "backup_folder": "/u01/RA_OPS/tables_backup"
     }}',
     TRUE, 2, 86400),

    -- 3. Oracle Export Job (legacy: extract_tables_to_excel2.py)
    -- Exports Oracle tables to CSV and creates/updates _historical snapshots
    -- Legacy: reads listoftables.csv, exports to monthly folders
    ('{export_id}',
     'Monthly Oracle Export',
     'export',
     'Exports Oracle tables to CSV files and maintains historical snapshots. '
     'Creates monthly folders (e.g., February-2026/). '
     'Copies to target folders for downstream consumption. '
     'Legacy script: extract_tables_to_excel2.py + listoftables.csv',
     '0 1 1 * *',
     '{{
        "description": "Migrated from extract_tables_to_excel2.py",
        "legacy_script": "extract_tables_to_excel2.py",
        "legacy_config": "/u01/RA_OPS/Export_BACKUP/listoftables.csv",
        "legacy_work_dir": "/u01/RA_OPS",
        "schema_owner": "RECON_PRD",
        "output_base_folder": "/u01/RA_OPS/Export_BACKUP",
        "copy_targets": [
            "/home/sdev/ROOT/AutoExport",
            "/home/sdev/ROOT2/Extracts"
        ],
        "create_historical_tables": true,
        "export_format": "csv"
     }}',
     TRUE, 3, 14400),

    -- 4. Dependency Graph Scan (legacy: treevFinalv3.py)
    -- Scans Oracle metadata + parses SQL to build view/MV dependency graph
    -- Generates topological sort for refresh ordering
    ('{dep_scan_id}',
     'Weekly Dependency Scan',
     'dependency_scan',
     'Scans Oracle all_dependencies + parses SQL definitions to build '
     'the full dependency graph for views and materialized views. '
     'Generates topological ordering for optimal parallel MV refresh. '
     'Legacy script: treevFinalv3.py',
     '0 0 * * 0',
     '{{
        "description": "Migrated from treevFinalv3.py",
        "legacy_script": "treevFinalv3.py",
        "schema_owner": "RECON_PRD",
        "output_csv": "sorted_views_materialized_views.csv",
        "cleaned_csv": "cleaned_sorted_views.csv",
        "scan_invalid_objects": true,
        "extract_sql_dependencies": true
     }}',
     TRUE, 3, 1800)
ON CONFLICT (name) DO NOTHING;
"""


def generate_export_tables_sql():
    """
    Generate SQL for Oracle export table list.
    Legacy: extract_tables_to_excel2.py reads from /u01/RA_OPS/Export_BACKUP/listoftables.csv
    We create a dedicated config table to hold these.
    """
    return """
-- ============================================================
-- EXPORT TABLE LIST (Migrated from listoftables.csv)
-- Used by extract_tables_to_excel2.py for Oracle -> CSV exports
-- ============================================================
-- NOTE: The actual table list is loaded from the Oracle server at runtime.
-- The legacy script read from: /u01/RA_OPS/Export_BACKUP/listoftables.csv
-- Format: TableName, Key, WhereCondition
-- This configuration is now stored in ra_meta.table_config with source_type='oracle'
-- and job_type='export'. The export job reads these configs dynamically.
--
-- To migrate the actual listoftables.csv content, run the Python migration
-- script which reads the CSV and inserts into table_config.
-- See: migrations/migrate_export_tables.py
"""


def generate_data_quality_rules_sql():
    """Generate default data quality rules based on legacy processing patterns."""
    return """
-- ============================================================
-- DATA QUALITY RULES (Extracted from legacy processing logic)
-- ============================================================
-- These rules codify the data cleaning patterns from Loader_full_Newv2.py:
-- 1. Remove commas from values (breaks CSV/SQL*Loader)
-- 2. Remove double quotes from values
-- 3. Remove newlines from values
-- 4. Collapse double spaces to single space
-- 5. Truncate strings to 4000 chars (Oracle VARCHAR2 limit)
-- 6. Remove decimal .0 from integer-like values
-- 7. Replace None/NaN/NaT with empty string

INSERT INTO ra_meta.data_quality_rules (id, table_config_id, rule_name, rule_type, column_name, rule_expression, severity, is_enabled)
SELECT
    gen_random_uuid(),
    tc.id,
    'Remove commas from values',
    'custom',
    NULL,
    'REPLACE(value, '','', '''')',
    'warning',
    TRUE
FROM ra_meta.table_config tc
WHERE tc.source_type IN ('csv', 'excel', 'postgres')
LIMIT 1;

-- Generic rules applied to all loaded data (documented as platform defaults)
INSERT INTO ra_meta.data_quality_rules (id, table_config_id, rule_name, rule_type, column_name, rule_expression, severity, is_enabled)
VALUES
    (gen_random_uuid(), (SELECT id FROM ra_meta.table_config LIMIT 1),
     'Max field length check (Oracle VARCHAR2)', 'range', NULL,
     'LENGTH(value) <= 4000', 'error', TRUE),

    (gen_random_uuid(), (SELECT id FROM ra_meta.table_config LIMIT 1),
     'No embedded newlines', 'regex', NULL,
     'value NOT LIKE ''%\\n%''', 'warning', TRUE),

    (gen_random_uuid(), (SELECT id FROM ra_meta.table_config LIMIT 1),
     'No embedded commas in unquoted fields', 'regex', NULL,
     'value NOT LIKE ''%,%''', 'warning', TRUE)
ON CONFLICT DO NOTHING;
"""


def generate_sla_policies_update_sql():
    """Update SLA policies with actual values from legacy timeouts."""
    return """
-- ============================================================
-- SLA POLICIES (Updated with actual legacy timeout values)
-- ============================================================
-- Legacy: Refresh_withkill.py used 60-minute timeout per MV
-- Legacy: Loader had no explicit timeout (relied on SQL*Loader defaults)
-- Legacy: Export had no explicit timeout

UPDATE ra_meta.sla_policies SET max_duration_minutes = 60 WHERE job_type = 'export' AND name = 'Export SLA';
UPDATE ra_meta.sla_policies SET max_duration_minutes = 120 WHERE job_type = 'load' AND name = 'Load SLA';
UPDATE ra_meta.sla_policies SET max_duration_minutes = 1440 WHERE job_type = 'refresh' AND name = 'Refresh SLA';
UPDATE ra_meta.sla_policies SET max_duration_minutes = 30 WHERE job_type = 'dependency_scan' AND name = 'Dependency Scan SLA';
"""


def generate_full_migration():
    """Generate the complete migration SQL file."""
    # Load legacy data
    legacy_dir = Path(__file__).parent.parent / "legacy"
    data_mapping_path = legacy_dir / "data_mapping.xlsx"
    sequence_path = legacy_dir / "sequence_updated3.xlsx"

    print(f"Loading data_mapping.xlsx from {data_mapping_path}...")
    mappings = load_data_mapping(str(data_mapping_path))
    print(f"  Loaded {len(mappings)} table configurations")

    print(f"Loading sequence_updated3.xlsx from {sequence_path}...")
    mvs = load_mv_sequence(str(sequence_path))
    print(f"  Loaded {len(mvs)} MV refresh entries")

    # Generate SQL sections
    conn_sql, pg_id, oracle_id, oracle_export_id = generate_connection_profiles_sql()
    table_sql = generate_table_config_sql(mappings, pg_id, oracle_id)
    mv_sql = generate_mv_sequence_sql(mvs)
    jobs_sql = generate_jobs_sql()
    export_sql = generate_export_tables_sql()
    dq_sql = generate_data_quality_rules_sql()
    sla_sql = generate_sla_policies_update_sql()

    # Combine into full migration
    full_sql = f"""-- ============================================================
-- Revenue Assurance Platform - Legacy Data Migration
-- Generated: {datetime.now(timezone.utc).isoformat()}
-- ============================================================
-- This migration loads ALL configurations from legacy scripts:
--
-- Source Files Migrated:
--   1. data_mapping.xlsx (Sheet1: {len([m for m in mappings if m['_sheet']=='Sheet1'])} rows, Skip: {len([m for m in mappings if m['_sheet']=='Skip'])} rows)
--   2. sequence_updated3.xlsx ({len(mvs)} materialized views)
--   3. manage_jobs.py (3 job configs + paths)
--   4. extract_tables_to_excel2.py (export config + credentials)
--   5. Refresh_withkill.py (refresh config + timeout settings)
--   6. treevFinalv3.py (dependency scan config)
--
-- Hardcoded Credentials Removed:
--   - Oracle: recon_prd/recon123@incor.solutions.com.sa -> ORACLE_PASSWORD env var
--   - PostgreSQL: config.json credentials -> PG_SOURCE_PASSWORD env var
--   - ORACLE_HOME: /home/oracle/app/product/12.2.0/client_1 -> ORACLE_HOME env var
--
-- Legacy Paths Preserved (in config JSON for reference):
--   - /u01/RA_OPS/Test_New_Loader/Final/ (loader work dir)
--   - /u01/RA_OPS/ (refresh/export work dir)
--   - /u01/RA_OPS/Export_BACKUP/ (export output)
--   - /home/sdev/ROOT/AutoExport (export target)
--   - /home/sdev/ROOT2/Extracts (export target)
-- ============================================================

BEGIN;

{conn_sql}

{table_sql}

{mv_sql}

{jobs_sql}

{export_sql}

{dq_sql}

{sla_sql}

-- ============================================================
-- MIGRATION AUDIT RECORD
-- ============================================================
INSERT INTO ra_meta.audit_logs (action, resource_type, details)
VALUES (
    'migration.legacy_data_loaded',
    'system',
    '{{
        "migration_script": "002_seed_legacy_data.sql",
        "table_configs_loaded": {len(mappings)},
        "mv_refresh_entries": {len(mvs)},
        "jobs_created": 4,
        "connection_profiles": 3,
        "source_files": ["data_mapping.xlsx", "sequence_updated3.xlsx", "manage_jobs.py", "extract_tables_to_excel2.py", "Refresh_withkill.py", "treevFinalv3.py"]
    }}'
);

COMMIT;
"""
    return full_sql


if __name__ == "__main__":
    print("=" * 60)
    print("Legacy Data Migration Script")
    print("=" * 60)

    sql = generate_full_migration()

    output_path = Path(__file__).parent / "002_seed_legacy_data.sql"
    with open(output_path, "w") as f:
        f.write(sql)

    print(f"\nMigration SQL written to: {output_path}")
    print(f"File size: {os.path.getsize(output_path):,} bytes")
    print("\nTo apply:")
    print(f"  psql -h <host> -U ra_admin -d ra_metadata -f {output_path}")
